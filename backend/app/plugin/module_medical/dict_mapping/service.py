"""医疗字典值映射 — 服务层。

核心职责：
- 映射规则的增删改查（参考 mapping_service.py 全量替换模式）
- 运行时归一化 normalize()：查 Redis Hash，命中返回标准值，未匹配写 unmatched
- 缓存刷新：写映射后刷新 Redis Hash
"""

import io
from datetime import datetime
from typing import Any

import pandas as pd
from fastapi import UploadFile
from redis.asyncio.client import Redis
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1.module_system.auth.schema import AuthSchema
from app.common.constant import PLATFORM_TENANT_ID
from app.common.enums import RedisInitKeyConfig
from app.core.database import async_db_session
from app.core.exceptions import CustomException
from app.core.logger import log
from app.core.redis_crud import RedisCURD

from .crud import DictMappingCRUD, DictUnmatchedCRUD
from .model import DictMappingModel, DictUnmatchedModel
from .schema import (
    DictMappingCreateSchema,
    DictMappingOutSchema,
    DictMappingUpdateSchema,
    NormalizeBatchIn,
    NormalizeIn,
    NormalizeResult,
)


def _cache_key(tenant_id: int, dict_type: str) -> str:
    """Redis Hash key：system_dict_mapping:{tenant_id}:{dict_type}"""
    return f"{RedisInitKeyConfig.SYSTEM_DICT_MAPPING.key}:{tenant_id}:{dict_type}"


class DictMappingService:
    """医疗字典值映射服务。"""

    @classmethod
    async def load_all_mappings(
        cls,
        db: AsyncSession,
        dict_type: str,
        hospital_id: int | None = None,
    ) -> dict[str, str]:
        """批量加载字典映射，供 ETL 在内存中归一化。"""
        from app.api.v1.module_system.dict.model import DictDataModel, DictTypeModel

        dt_result = await db.execute(
            select(DictTypeModel).where(DictTypeModel.dict_type == dict_type)
        )
        dt_obj = dt_result.scalars().first()
        if not dt_obj:
            log.warning("load_all_mappings: dict_type=%s 不存在", dict_type)
            return {}

        stmt = (
            select(DictMappingModel.raw_label, DictDataModel.dict_value)
            .join(DictDataModel, DictMappingModel.dict_data_id == DictDataModel.id)
            .where(DictMappingModel.dict_type_id == dt_obj.id)
        )
        if hospital_id is not None:
            stmt = stmt.where(
                DictMappingModel.hospital_id.in_([hospital_id, PLATFORM_TENANT_ID])
            )
        rows = (await db.execute(stmt)).all()
        return {str(raw).strip().lower(): value for raw, value in rows if raw and value}


    @classmethod
    async def list_service(
        cls,
        auth: AuthSchema,
        hospital_id: int | None = None,
        dict_type_id: int | None = None,
        raw_label: str | None = None,
    ) -> list[dict]:
        """列出映射规则。"""
        search: dict[str, Any] = {}
        if hospital_id is not None:
            search["hospital_id"] = hospital_id
        if dict_type_id is not None:
            search["dict_type_id"] = dict_type_id
        if raw_label is not None:
            search["raw_label"] = raw_label

        items = await DictMappingCRUD(auth).list(search=search)
        return [DictMappingOutSchema.model_validate(obj).model_dump() for obj in items]

    @classmethod
    async def get_detail_service(cls, auth: AuthSchema, id: int) -> dict:
        """获取单条映射详情。"""
        obj = await DictMappingCRUD(auth).get(id=id)
        if not obj:
            raise CustomException(msg="映射规则不存在")
        return DictMappingOutSchema.model_validate(obj).model_dump()

    @classmethod
    async def create_service(
        cls, auth: AuthSchema, redis: Redis, data: DictMappingCreateSchema
    ) -> dict:
        """创建映射规则。"""
        # 唯一性检查（大小写归一）
        exist = await DictMappingCRUD(auth).get_by_raw_label(
            hospital_id=data.hospital_id,
            dict_type_id=data.dict_type_id,
            raw_label=data.raw_label,
        )
        if exist:
            raise CustomException(msg="该映射已存在（同一医院同一类型下 raw_label 重复）")

        obj = await DictMappingCRUD(auth).create(data=data)
        result = DictMappingOutSchema.model_validate(obj).model_dump()

        # 刷新缓存（传入 auth.db 以读取已 flush 的未提交行）
        await cls._refresh_cache(redis, auth.user.tenant_id, data.dict_type_id, db=auth.db)
        return result

    @classmethod
    async def update_service(
        cls,
        auth: AuthSchema,
        redis: Redis,
        id: int,
        data: DictMappingUpdateSchema,
    ) -> dict:
        """更新映射规则。"""
        obj = await DictMappingCRUD(auth).get(id=id)
        if not obj:
            raise CustomException(msg="映射规则不存在")

        updated = await DictMappingCRUD(auth).update(id=id, data=data)
        result = DictMappingOutSchema.model_validate(updated).model_dump()

        # 刷新缓存（用更新后对象的 dict_type_id，因为部分更新可能未传该字段）
        await cls._refresh_cache(redis, auth.user.tenant_id, updated.dict_type_id, db=auth.db)
        return result

    @classmethod
    async def delete_service(
        cls, auth: AuthSchema, redis: Redis, ids: list[int]
    ) -> None:
        """删除映射规则。"""
        # 先获取 dict_type_id 用于刷新缓存
        dict_type_ids: set[int] = set()
        for id_ in ids:
            obj = await DictMappingCRUD(auth).get(id=id_)
            if obj:
                dict_type_ids.add(obj.dict_type_id)

        await DictMappingCRUD(auth).delete(ids=ids)

        # 刷新缓存（传入 auth.db 以读取已 flush 的未提交行）
        for dtid in dict_type_ids:
            await cls._refresh_cache(redis, auth.user.tenant_id, dtid, db=auth.db)

    @classmethod
    async def batch_create_service(
        cls,
        auth: AuthSchema,
        redis: Redis,
        items: list[DictMappingCreateSchema],
    ) -> list[dict]:
        """批量创建映射规则（全量替换模式）。"""
        if not items:
            raise CustomException(msg="批量创建列表不能为空")

        # 按 (hospital_id, dict_type_id) 分组
        groups: dict[tuple[int, int], list[DictMappingCreateSchema]] = {}
        for item in items:
            key = (item.hospital_id, item.dict_type_id)
            groups.setdefault(key, []).append(item)

        results = []
        for (hospital_id, dict_type_id), group_items in groups.items():
            # 删除旧映射
            old_items = await DictMappingCRUD(auth).list(
                search={"hospital_id": hospital_id, "dict_type_id": dict_type_id}
            )
            if old_items:
                await DictMappingCRUD(auth).delete(ids=[o.id for o in old_items])

            # 批量新建
            for item in group_items:
                obj = await DictMappingCRUD(auth).create(data=item)
                results.append(DictMappingOutSchema.model_validate(obj).model_dump())

            # 刷新缓存（传入 auth.db 以读取已 flush 的未提交行）
            await cls._refresh_cache(redis, auth.user.tenant_id, dict_type_id, db=auth.db)

        return results

    # ------------------------------------------------------------------
    # 运行时归一化
    # ------------------------------------------------------------------

    @classmethod
    async def normalize_service(
        cls,
        auth: AuthSchema,
        redis: Redis,
        data: NormalizeIn,
    ) -> NormalizeResult:
        """单值归一化：查 Redis Hash，命中返回标准值，未匹配写 unmatched。"""
        from sqlalchemy.ext.asyncio import AsyncSession

        db: AsyncSession = auth.db

        # 先查 dict_type_id
        dict_type_id = await cls._get_dict_type_id(db, data.dict_type)
        if dict_type_id is None:
            return NormalizeResult(raw_label=data.raw_label, matched=False)

        tenant_id = auth.user.tenant_id if auth.user else PLATFORM_TENANT_ID

        # 查缓存
        cached = await cls._lookup_cache(redis, tenant_id, data.dict_type, data.raw_label)
        if cached is not None:
            return NormalizeResult(raw_label=data.raw_label, dict_value=cached, matched=True)

        # 缓存未命中：回源 DB 查映射
        mapping = await DictMappingCRUD(auth).get_by_raw_label(
            hospital_id=data.hospital_id,
            dict_type_id=dict_type_id,
            raw_label=data.raw_label,
        )
        if mapping and mapping.dict_data_id:
            # 查 dict_value
            dict_value = await cls._get_dict_value(db, mapping.dict_data_id)
            if dict_value:
                # 回填缓存
                await cls._set_cache(
                    redis, tenant_id, data.dict_type, data.raw_label, dict_value
                )
                return NormalizeResult(
                    raw_label=data.raw_label, dict_value=dict_value, matched=True
                )

        # 未匹配：写 unmatched（独立 session，不影响请求事务）
        async with async_db_session() as new_db:
            await DictUnmatchedCRUD(auth).upsert_unmatched(
                db=new_db,
                hospital_id=data.hospital_id,
                dict_type_id=dict_type_id,
                raw_label=data.raw_label,
                raw_value=None,
                tenant_id=tenant_id,
            )
            await new_db.commit()

        return NormalizeResult(raw_label=data.raw_label, matched=False)

    @classmethod
    async def normalize_batch_service(
        cls,
        auth: AuthSchema,
        redis: Redis,
        data: NormalizeBatchIn,
    ) -> list[NormalizeResult]:
        """批量归一化。"""
        results = []
        for raw_label in data.raw_labels:
            result = await cls.normalize_service(
                auth=auth,
                redis=redis,
                data=NormalizeIn(
                    hospital_id=data.hospital_id,
                    dict_type=data.dict_type,
                    raw_label=raw_label,
                ),
            )
            results.append(result)
        return results

    # ------------------------------------------------------------------
    # 未匹配记录
    # ------------------------------------------------------------------

    @classmethod
    async def list_unmatched_service(
        cls,
        auth: AuthSchema,
        hospital_id: int | None = None,
        dict_type_id: int | None = None,
    ) -> list[dict]:
        """列出未匹配记录。"""
        search: dict[str, Any] = {}
        if hospital_id is not None:
            search["hospital_id"] = hospital_id
        if dict_type_id is not None:
            search["dict_type_id"] = dict_type_id

        items = await DictUnmatchedCRUD(auth).list(search=search)
        return [_unmatched_to_dict(obj) for obj in items]

    @classmethod
    async def resolve_unmatched_service(
        cls,
        auth: AuthSchema,
        redis: Redis,
        unmatched_id: int,
        mapping_id: int | None = None,
    ) -> None:
        """解决未匹配记录。"""
        obj = await DictUnmatchedCRUD(auth).get(id=unmatched_id)
        if not obj:
            raise CustomException(msg="未匹配记录不存在")

        obj.resolution = "resolve" if mapping_id else "ignore"
        user_id = auth.user.id if auth.user else None
        tenant_id = auth.user.tenant_id if auth.user else PLATFORM_TENANT_ID
        obj.resolved_by = user_id
        obj.resolved_at = datetime.now()
        obj.resolved_as_mapping_id = mapping_id
        obj.status = "2" if mapping_id else "1"
        await auth.db.flush()

        # 如果关联了映射，刷新缓存（传入 auth.db 以读取已 flush 的未提交行）
        if mapping_id:
            await cls._refresh_cache(redis, tenant_id, obj.dict_type_id, db=auth.db)

    # ------------------------------------------------------------------
    # 缓存
    # ------------------------------------------------------------------

    @classmethod
    async def refresh_cache_service(
        cls, auth: AuthSchema, redis: Redis, dict_type_id: int
    ) -> None:
        """手动刷新某类型的 Redis 缓存。"""
        tenant_id = auth.user.tenant_id if auth.user else PLATFORM_TENANT_ID
        await cls._refresh_cache(redis, tenant_id, dict_type_id)

    @classmethod
    async def _refresh_cache(
        cls,
        redis: Redis,
        tenant_id: int,
        dict_type_id: int,
        db: AsyncSession | None = None,
    ) -> None:
        """从 DB 加载某类型的所有映射到 Redis Hash。

        db 非空时直接用（请求事务已 flush，可见未提交行）；
        db 为空时开独立 session（手动刷新场景）。
        """
        async def _do_refresh(session: AsyncSession) -> tuple[str | None, dict[str, str]]:
            from app.api.v1.module_system.dict.model import DictDataModel, DictTypeModel

            # 查 dict_type 名
            dt_result = await session.execute(
                select(DictTypeModel).where(DictTypeModel.id == dict_type_id)
            )
            dict_type_obj = dt_result.scalars().first()
            if not dict_type_obj:
                log.warning("字典映射缓存刷新: dict_type_id=%s 不存在", dict_type_id)
                return None, {}

            dict_type_name = dict_type_obj.dict_type

            # JOIN 批量查映射 + 字典值，消除 N+1
            rows = (
                await session.execute(
                    select(DictMappingModel.raw_label, DictDataModel.dict_value)
                    .join(
                        DictDataModel,
                        DictMappingModel.dict_data_id == DictDataModel.id,
                    )
                    .where(DictMappingModel.dict_type_id == dict_type_id)
                )
            ).all()

            hash_data: dict[str, str] = {
                r[0].lower(): r[1] for r in rows if r[1]
            }
            return dict_type_name, hash_data

        try:
            if db is not None:
                dict_type_name, hash_data = await _do_refresh(db)
            else:
                async with async_db_session() as new_db:
                    dict_type_name, hash_data = await _do_refresh(new_db)

            if dict_type_name is None:
                return

            cache_k = _cache_key(tenant_id, dict_type_name)
            if hash_data:
                # 批量写入 Hash：用 redis.hset(mapping=...) 一次写入全部字段
                await redis.hset(name=cache_k, mapping=hash_data)
            else:
                # 无映射则删除 key
                await RedisCURD(redis).delete(cache_k)

            log.info(
                "字典映射缓存刷新: tenant=%s dict_type=%s count=%d",
                tenant_id, dict_type_name, len(hash_data),
            )
        except Exception as e:
            log.error("字典映射缓存刷新失败: %s", e)

    @classmethod
    async def _lookup_cache(
        cls, redis: Redis, tenant_id: int, dict_type: str, raw_label: str
    ) -> str | None:
        """从 Redis Hash 查单个映射。"""
        try:
            cache_k = _cache_key(tenant_id, dict_type)
            values = await RedisCURD(redis).hash_get(name=cache_k, keys=[raw_label.lower()])
            if values and values[0]:
                v = values[0]
                return v if isinstance(v, str) else v.decode() if isinstance(v, bytes) else str(v)
        except Exception:
            pass
        return None

    @classmethod
    async def _set_cache(
        cls, redis: Redis, tenant_id: int, dict_type: str, raw_label: str, dict_value: str
    ) -> None:
        """写入 Redis Hash 单个映射。"""
        try:
            cache_k = _cache_key(tenant_id, dict_type)
            await RedisCURD(redis).hash_set(name=cache_k, key=raw_label.lower(), value=dict_value)
        except Exception as e:
            log.warning("字典映射缓存写入失败: %s", e)

    # ------------------------------------------------------------------
    # Excel 批量导入
    # ------------------------------------------------------------------

    @classmethod
    async def import_excel_service(
        cls, auth: AuthSchema, redis: Redis, file: UploadFile
    ) -> str:
        """Excel 批量导入映射规则（覆盖更新模式）。

        Excel 格式：每个 sheet 对应一家医院，sheet 名 = center_code。
        每个 sheet 三列：dict_type | raw_label | dict_value。
        遇到已存在的 (hospital_id, dict_type_id, raw_label) 覆盖更新 dict_data_id。
        """
        from app.plugin.module_medical.hospital.model import HospitalModel

        try:
            contents = await file.read()
            await file.close()
        except Exception as e:
            raise CustomException(msg=f"读取文件失败: {e!s}")

        # 读取所有 sheet → {sheet_name: DataFrame}
        try:
            sheets = pd.read_excel(io.BytesIO(contents), sheet_name=None, dtype=str)
        except Exception as e:
            raise CustomException(msg=f"解析 Excel 失败: {e!s}")

        if not sheets:
            raise CustomException(msg="Excel 文件不包含任何 sheet")

        # 缓存：center_code → hospital_id、dict_type → dict_type_id、(dict_type_id, dict_value) → dict_data_id
        hospital_cache: dict[str, int | None] = {}
        dict_type_cache: dict[str, int | None] = {}
        dict_data_cache: dict[tuple[int, str], int | None] = {}
        refreshed_dict_types: set[int] = set()

        crud = DictMappingCRUD(auth)
        db = auth.db
        success_count = 0
        sheet_count = 0
        error_msgs: list[str] = []

        for sheet_name, df in sheets.items():
            center_code = str(sheet_name).strip()
            if df is None or df.empty:
                continue

            # 校验表头
            expected = {"dict_type", "raw_label", "dict_value"}
            actual = {str(c).strip() for c in df.columns}
            if not expected.issubset(actual):
                error_msgs.append(f"sheet[{center_code}]: 表头缺失，需含 {expected}")
                continue

            # center_code → hospital_id
            if center_code not in hospital_cache:
                sql = select(HospitalModel.id).where(HospitalModel.code == center_code)
                result = await db.execute(sql)
                hospital_cache[center_code] = result.scalars().first()

            hospital_id = hospital_cache[center_code]
            if hospital_id is None:
                error_msgs.append(
                    f"sheet[{center_code}]: 未找到 code={center_code} 的医院，跳过该 sheet"
                )
                continue

            sheet_count += 1

            # 去掉列名空白
            df.columns = [str(c).strip() for c in df.columns]
            for idx, row in df.iterrows():
                excel_row = int(idx) + 2  # +1 表头, +1 从1计数
                dict_type = str(row.get("dict_type", "") or "").strip()
                raw_label = str(row.get("raw_label", "") or "").strip()
                dict_value = str(row.get("dict_value", "") or "").strip()

                if not raw_label:
                    error_msgs.append(
                        f"sheet[{center_code}]第{excel_row}行: raw_label 为空，跳过"
                    )
                    continue

                # dict_type → dict_type_id
                if dict_type not in dict_type_cache:
                    dict_type_cache[dict_type] = await cls._get_dict_type_id(db, dict_type)
                dict_type_id = dict_type_cache[dict_type]
                if dict_type_id is None:
                    error_msgs.append(
                        f"sheet[{center_code}]第{excel_row}行: dict_type={dict_type} 不存在，跳过"
                    )
                    continue

                # dict_value → dict_data_id
                data_key = (dict_type_id, dict_value)
                if data_key not in dict_data_cache:
                    dict_data_cache[data_key] = await cls._get_dict_data_id_by_value(
                        db, dict_type_id, dict_value
                    )
                dict_data_id = dict_data_cache[data_key]
                if dict_data_id is None:
                    error_msgs.append(
                        f"sheet[{center_code}]第{excel_row}行: dict_value={dict_value} "
                        f"在 {dict_type} 下不存在，跳过"
                    )
                    continue

                # upsert（覆盖更新）
                existing = await crud.get_by_raw_label(hospital_id, dict_type_id, raw_label)
                if existing:
                    if existing.dict_data_id != dict_data_id:
                        await crud.update(
                            id=existing.id,
                            data=DictMappingUpdateSchema(dict_data_id=dict_data_id),
                        )
                    success_count += 1
                else:
                    await crud.create(
                        data=DictMappingCreateSchema(
                            hospital_id=hospital_id,
                            dict_type_id=dict_type_id,
                            dict_data_id=dict_data_id,
                            raw_label=raw_label,
                        )
                    )
                    success_count += 1

                # 收集需要刷新缓存的 dict_type_id
                refreshed_dict_types.add(dict_type_id)

        # 刷新所有涉及类型的缓存
        tenant_id = auth.user.tenant_id if auth.user else PLATFORM_TENANT_ID
        for dt_id in refreshed_dict_types:
            await cls._refresh_cache(redis, tenant_id, dt_id, db=db)

        result = f"成功导入 {success_count} 条映射（{sheet_count} 个医院）"
        if error_msgs:
            result += f"；跳过 {len(error_msgs)} 条无效数据"
        return result

    @classmethod
    async def get_import_template_service(cls) -> bytes:
        """生成 Excel 导入模板（单 sheet 示例 + 表头说明）。"""
        from app.utils.excel_util import ExcelUtil

        # 示例数据：一个名为 "示例_center_code" 的 sheet
        example_df = pd.DataFrame(
            [
                {"dict_type": "med_sex", "raw_label": "男", "dict_value": "M"},
                {"dict_type": "med_sex", "raw_label": "m", "dict_value": "M"},
                {"dict_type": "med_sex", "raw_label": "女", "dict_value": "F"},
                {"dict_type": "med_laterality", "raw_label": "左", "dict_value": "L"},
                {"dict_type": "med_laterality", "raw_label": "右", "dict_value": "R"},
            ]
        )

        # 用 ExcelUtil 生成带表头的模板
        header_list = ["dict_type", "raw_label", "dict_value"]
        base_bytes = ExcelUtil.get_excel_template(
            header_list=header_list,
            selector_header_list=[],
            option_list=[],
        )

        # 重新打开，把示例数据写入名为 "示例_center_code" 的 sheet
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(base_bytes))
        # ExcelUtil 生成的模板默认有一个 sheet，重命名为示例名并写入示例数据
        ws = wb.active
        ws.title = "示例_center_code"
        for r_idx, row in enumerate(example_df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    @classmethod
    async def export_excel_service(cls, auth: AuthSchema) -> bytes:
        """导出全部选项映射到 Excel（每家医院一个 sheet，可直接用于导入）。

        导出格式与导入模板完全一致：
        - 每个 sheet = 一家医院，sheet 名 = center_code (= HospitalModel.code)
        - 三列：dict_type | raw_label | dict_value
        - dict_data_id 反查 sys_dict_data.dict_value（字符串）
        """
        from openpyxl import Workbook

        from app.api.v1.module_system.dict.model import DictDataModel, DictTypeModel
        from app.plugin.module_medical.hospital.model import HospitalModel

        db = auth.db

        # 一次性查出全部映射 + 关联的 dict_type / dict_data
        sql = (
            select(
                DictMappingModel.hospital_id,
                DictMappingModel.dict_type_id,
                DictMappingModel.raw_label,
                DictTypeModel.dict_type.label("dict_type_name"),
                DictDataModel.dict_value.label("dict_value"),
            )
            .outerjoin(DictTypeModel, DictMappingModel.dict_type_id == DictTypeModel.id)
            .outerjoin(DictDataModel, DictMappingModel.dict_data_id == DictDataModel.id)
            .order_by(DictMappingModel.hospital_id, DictMappingModel.dict_type_id)
        )
        sql = await DictMappingCRUD(auth)._CRUDBase__filter_permissions(sql)
        result = await db.execute(sql)
        rows = result.all()

        if not rows:
            raise CustomException(msg="当前没有可导出的映射数据")

        # 预加载 hospital_id → center_code 映射
        hospital_ids = {r.hospital_id for r in rows}
        code_sql = select(HospitalModel.id, HospitalModel.code).where(
            HospitalModel.id.in_(hospital_ids)
        )
        code_result = await db.execute(code_sql)
        code_map = {r.id: r.code for r in code_result.all()}

        # 按 hospital_id 分组
        groups: dict[int, list] = {}
        for r in rows:
            groups.setdefault(r.hospital_id, []).append(r)

        # 生成 Excel（每家医院一个 sheet）
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认空 sheet

        for hospital_id, group_rows in groups.items():
            center_code = code_map.get(hospital_id, f"hospital_{hospital_id}")
            # openpyxl sheet 名最长 31 字符，且不能含 []:*?/\\
            sheet_name = center_code[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["dict_type", "raw_label", "dict_value"])
            for r in group_rows:
                ws.append(
                    [
                        r.dict_type_name or "",
                        r.raw_label or "",
                        r.dict_value or "",
                    ]
                )

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    @classmethod
    async def _get_hospital_id_by_code(
        cls, db: AsyncSession, center_code: str
    ) -> int | None:
        """按 center_code (= HospitalModel.code) 查 hospital_id。"""
        try:
            from app.plugin.module_medical.hospital.model import HospitalModel

            sql = select(HospitalModel.id).where(HospitalModel.code == center_code)
            result = await db.execute(sql)
            return result.scalars().first()
        except Exception:
            return None

    @classmethod
    async def _get_dict_data_id_by_value(
        cls, db: AsyncSession, dict_type_id: int, dict_value: str
    ) -> int | None:
        """按 (dict_type_id, dict_value) 查 sys_dict_data.id。"""
        try:
            from app.api.v1.module_system.dict.model import DictDataModel

            sql = select(DictDataModel.id).where(
                DictDataModel.dict_type_id == dict_type_id,
                DictDataModel.dict_value == dict_value,
            )
            result = await db.execute(sql)
            return result.scalars().first()
        except Exception:
            return None

    @classmethod
    async def _get_dict_type_id(cls, db: AsyncSession, dict_type: str) -> int | None:
        """按 dict_type 字符串查 id。"""
        try:
            from app.api.v1.module_system.dict.model import DictTypeModel

            sql = select(DictTypeModel).where(DictTypeModel.dict_type == dict_type)
            result = await db.execute(sql)
            obj = result.scalars().first()
            return obj.id if obj else None
        except Exception:
            return None

    @classmethod
    async def _get_dict_value(cls, db: AsyncSession, dict_data_id: int) -> str | None:
        """按 dict_data_id 查 dict_value。"""
        try:
            from app.api.v1.module_system.dict.model import DictDataModel

            sql = select(DictDataModel).where(DictDataModel.id == dict_data_id)
            result = await db.execute(sql)
            obj = result.scalars().first()
            return obj.dict_value if obj else None
        except Exception:
            return None


def _unmatched_to_dict(obj: DictUnmatchedModel) -> dict:
    """未匹配记录转字典。"""
    dict_type_name = None
    if hasattr(obj, "dict_type") and obj.dict_type:
        dict_type_name = getattr(obj.dict_type, "dict_name", None)

    return {
        "id": obj.id,
        "hospital_id": obj.hospital_id,
        "dict_type_id": obj.dict_type_id,
        "dict_type_name": dict_type_name,
        "raw_label": obj.raw_label,
        "raw_value": obj.raw_value,
        "occurrence_count": obj.occurrence_count,
        "last_seen_at": obj.last_seen_at,
        "resolution": obj.resolution,
    }
